from fastapi import FastAPI, Depends, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
import os
import hashlib
import secrets

from database import engine, get_db, Base
from models import Teacher, Block, SchoolClass, Timetable

Base.metadata.create_all(bind=engine)

app = FastAPI(title="KHMHS Alathiyur - Timetable")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ===== Authentication =====
USERS = {
    "admin": {"password": hashlib.sha256("khmhs2026".encode()).hexdigest(), "role": "admin"},
    "staff": {"password": hashlib.sha256("timetable123".encode()).hexdigest(), "role": "staff"},
}
active_tokens = {}  # token -> {"username": ..., "role": ...}


class LoginRequest(BaseModel):
    username: str
    password: str


@app.post("/api/login")
def login(req: LoginRequest):
    pw_hash = hashlib.sha256(req.password.encode()).hexdigest()
    user = USERS.get(req.username)
    if user and user["password"] == pw_hash:
        token = secrets.token_hex(32)
        active_tokens[token] = {"username": req.username, "role": user["role"]}
        return {"token": token, "username": req.username, "role": user["role"]}
    raise HTTPException(401, "Invalid credentials")


@app.get("/api/verify")
def verify(request: Request):
    auth = request.headers.get("Authorization", "")
    token = auth.replace("Bearer ", "")
    if token in active_tokens:
        info = active_tokens[token]
        return {"ok": True, "username": info["username"], "role": info["role"]}
    raise HTTPException(401, "Invalid token")


@app.post("/api/logout")
def logout(request: Request):
    auth = request.headers.get("Authorization", "")
    token = auth.replace("Bearer ", "")
    active_tokens.pop(token, None)
    return {"ok": True}


def get_current_role(request: Request):
    """Helper to get current user's role"""
    auth = request.headers.get("Authorization", "")
    token = auth.replace("Bearer ", "")
    info = active_tokens.get(token)
    return info["role"] if info else None


# ===== Schemas =====
class TeacherSchema(BaseModel):
    name: Optional[str] = None
    maxPeriodsPerDay: Optional[int] = 7
    isBlockHead: Optional[bool] = False
    headOfBlock: Optional[str] = ""


class BlockSchema(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = ""
    head: Optional[str] = ""


class ClassSchema(BaseModel):
    name: Optional[str] = None
    divisions: Optional[list] = None
    block: Optional[str] = ""
    classType: Optional[str] = ""
    classTeacher: Optional[str] = ""
    subjects: Optional[list] = None


# ===== Teachers =====
@app.get("/api/teachers")
def get_teachers(db: Session = Depends(get_db)):
    teachers = db.query(Teacher).order_by(Teacher.name).all()
    return [{"id": t.id, "name": t.name, "maxPeriodsPerDay": t.maxPeriodsPerDay,
             "isBlockHead": t.isBlockHead, "headOfBlock": t.headOfBlock} for t in teachers]


@app.post("/api/teachers")
def create_teacher(data: TeacherSchema, db: Session = Depends(get_db)):
    t = Teacher(name=data.name, maxPeriodsPerDay=data.maxPeriodsPerDay,
                isBlockHead=data.isBlockHead, headOfBlock=data.headOfBlock)
    db.add(t)
    db.commit()
    db.refresh(t)
    return {"id": t.id, "name": t.name}


@app.put("/api/teachers/{tid}")
def update_teacher(tid: int, data: TeacherSchema, db: Session = Depends(get_db)):
    t = db.query(Teacher).filter(Teacher.id == tid).first()
    if not t:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(t, k, v)
    db.commit()
    return {"ok": True}


@app.delete("/api/teachers/{tid}")
def delete_teacher(tid: int, db: Session = Depends(get_db)):
    t = db.query(Teacher).filter(Teacher.id == tid).first()
    if t:
        db.delete(t)
        db.commit()
    return {"ok": True}


@app.put("/api/teachers/bulk-update-max-periods")
def bulk_update_max_periods(data: dict, db: Session = Depends(get_db)):
    max_periods = data.get("maxPeriodsPerDay", 5)
    db.query(Teacher).update({Teacher.maxPeriodsPerDay: max_periods})
    db.commit()
    return {"ok": True, "updated": "all"}


# ===== Blocks =====
@app.get("/api/blocks")
def get_blocks(db: Session = Depends(get_db)):
    blocks = db.query(Block).all()
    return [{"id": b.id, "name": b.name, "description": b.description, "head": b.head} for b in blocks]


@app.post("/api/blocks")
def create_block(data: BlockSchema, db: Session = Depends(get_db)):
    b = Block(name=data.name, description=data.description, head=data.head)
    db.add(b)
    db.commit()
    db.refresh(b)
    return {"id": b.id, "name": b.name}


@app.put("/api/blocks/{bid}")
def update_block(bid: int, data: BlockSchema, db: Session = Depends(get_db)):
    b = db.query(Block).filter(Block.id == bid).first()
    if not b:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(b, k, v)
    db.commit()
    return {"ok": True}


@app.delete("/api/blocks/{bid}")
def delete_block(bid: int, db: Session = Depends(get_db)):
    b = db.query(Block).filter(Block.id == bid).first()
    if b:
        db.delete(b)
        db.commit()
    return {"ok": True}


# ===== Classes =====
@app.get("/api/classes")
def get_classes(db: Session = Depends(get_db)):
    classes = db.query(SchoolClass).all()
    return [{"id": c.id, "name": c.name, "divisions": c.divisions, "block": c.block,
             "classType": c.classType, "classTeacher": c.classTeacher, "subjects": c.subjects} for c in classes]


@app.post("/api/classes")
def create_class(data: ClassSchema, db: Session = Depends(get_db)):
    c = SchoolClass(name=data.name, divisions=data.divisions, block=data.block,
                    classType=data.classType, classTeacher=data.classTeacher, subjects=data.subjects)
    db.add(c)
    db.commit()
    db.refresh(c)
    return {"id": c.id, "name": c.name}


@app.put("/api/classes/{cid}")
def update_class(cid: int, data: ClassSchema, db: Session = Depends(get_db)):
    c = db.query(SchoolClass).filter(SchoolClass.id == cid).first()
    if not c:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(c, k, v)
    db.commit()
    return {"ok": True}


@app.delete("/api/classes/{cid}")
def delete_class(cid: int, db: Session = Depends(get_db)):
    c = db.query(SchoolClass).filter(SchoolClass.id == cid).first()
    if c:
        db.delete(c)
        db.commit()
    return {"ok": True}


@app.delete("/api/classes")
def delete_classes_bulk(db: Session = Depends(get_db), className: str = None, block: str = None):
    query = db.query(SchoolClass)
    if className:
        query = query.filter(SchoolClass.name == className)
    if block:
        query = query.filter(SchoolClass.block == block)
    count = query.delete(synchronize_session=False)
    db.commit()
    return {"ok": True, "deleted": count}


# ===== Timetable =====
@app.get("/api/timetable")
def get_timetable(db: Session = Depends(get_db)):
    # Try id=1 first (convention for "current")
    tt = db.query(Timetable).filter(Timetable.id == 1).first()
    if not tt:
        # Fallback: get the latest entry that looks like a timetable
        tt = db.query(Timetable).order_by(Timetable.id.desc()).first()
    if not tt:
        return {}
    data = tt.data
    if not data:
        return {}
    # Unwrap if stored as {"timetable": {...}, "saved_at": ...}
    if isinstance(data, dict) and "timetable" in data and isinstance(data["timetable"], dict):
        return data["timetable"]
    # Check if it has class division keys (like "10-EA") - that means it's raw timetable
    if isinstance(data, dict):
        # Remove internal keys
        return {k: v for k, v in data.items() if not k.startswith('_')}
    return {}


@app.post("/api/timetable")
def save_timetable(data: dict, db: Session = Depends(get_db)):
    # Save to history first
    from datetime import datetime
    db.execute(
        Timetable.__table__.insert().values(data={"timetable": data, "saved_at": datetime.utcnow().isoformat()})
    )
    # Update current (id=1 is always current)
    current = db.query(Timetable).filter(Timetable.id == 1).first()
    if current:
        current.data = data
    else:
        db.add(Timetable(data=data))
    db.commit()
    return {"ok": True}


@app.get("/api/debug-solver")
def debug_solver(db: Session = Depends(get_db)):
    """Debug endpoint - shows what data the solver would receive"""
    classes = db.query(SchoolClass).all()
    teachers = db.query(Teacher).all()

    issues = []
    class_summaries = []

    for cls in classes:
        total_periods = 0
        missing_teachers = []
        for sub in (cls.subjects or []):
            periods = sub.get('periodsPerWeek', 0)
            if periods > 0:
                total_periods += periods
                if not sub.get('teacher'):
                    missing_teachers.append(sub.get('name', '?'))

        class_summaries.append({
            "name": cls.name,
            "divisions": cls.divisions,
            "totalPeriods": total_periods,
            "subjectCount": len([s for s in (cls.subjects or []) if s.get('periodsPerWeek', 0) > 0]),
            "missingTeachers": missing_teachers
        })

        if total_periods == 0:
            issues.append(f"Class {cls.name} ({cls.divisions}): NO periods assigned")
        elif total_periods > 35:
            issues.append(f"Class {cls.name} ({cls.divisions}): {total_periods} periods (exceeds 35)")
        if missing_teachers:
            issues.append(f"Class {cls.name}: subjects without teachers: {missing_teachers}")

    # Check teacher workload
    teacher_workload = {}
    for cls in classes:
        for div in (cls.divisions or []):
            for sub in (cls.subjects or []):
                if sub.get('teacher') and sub.get('periodsPerWeek', 0) > 0:
                    t = sub['teacher']
                    teacher_workload[t] = teacher_workload.get(t, 0) + sub['periodsPerWeek']

    overloaded = {t: w for t, w in teacher_workload.items() if w > 35}
    if overloaded:
        for t, w in overloaded.items():
            issues.append(f"Teacher '{t}': {w} total periods/week (max possible is 35)")

    return {
        "totalClasses": len(classes),
        "totalTeachers": len(teachers),
        "totalDivisions": sum(len(c.divisions or []) for c in classes),
        "classSummaries": class_summaries,
        "teacherWorkload": teacher_workload,
        "issues": issues,
        "status": "OK - ready to solve" if not issues else f"{len(issues)} issue(s) found"
    }


@app.post("/api/generate-timetable")
def generate_timetable(db: Session = Depends(get_db)):
    """Generate timetable using greedy solver"""
    from datetime import datetime
    import traceback

    # Get all data
    classes = db.query(SchoolClass).all()
    teachers = db.query(Teacher).all()

    classes_data = [
        {
            "name": c.name,
            "divisions": c.divisions,
            "block": c.block,
            "classType": c.classType,
            "classTeacher": c.classTeacher,
            "subjects": c.subjects
        }
        for c in classes
    ]

    teachers_data = [
        {
            "name": t.name,
            "maxPeriodsPerDay": t.maxPeriodsPerDay,
            "isBlockHead": t.isBlockHead,
            "headOfBlock": t.headOfBlock
        }
        for t in teachers
    ]

    # Validate data before solving
    if not classes_data:
        raise HTTPException(400, "No classes found. Please add classes first.")
    if not teachers_data:
        raise HTTPException(400, "No teachers found. Please add teachers first.")

    # Check that classes have subjects with teachers assigned
    empty_classes = [c['name'] for c in classes_data
                     if not any(s.get('periodsPerWeek', 0) > 0 and s.get('teacher')
                                for s in c.get('subjects', []))]
    if empty_classes:
        raise HTTPException(400, f"Classes with no teacher assignments: {', '.join(set(empty_classes))}. "
                                 "Please assign teachers to subjects first.")

    try:
        from solver import solve_timetable
        timetable = solve_timetable(classes_data, teachers_data, max_attempts=80)
    except (ValueError, RuntimeError) as e:
        print(f"Solver failed: {str(e)}")
        raise HTTPException(500, f"Solver failed: {str(e)}")
    except Exception as e:
        error_detail = traceback.format_exc()
        print(f"Solver error: {error_detail}")
        raise HTTPException(500, f"Solver error: {str(e)}")

    if timetable is None:
        raise HTTPException(500, "Solver returned no solution. Please check subject/teacher assignments.")

    # Save to history
    try:
        db.execute(
            Timetable.__table__.insert().values(data={"timetable": timetable, "saved_at": datetime.utcnow().isoformat()})
        )
        # Save as current (id=1 is the "current" slot)
        current = db.query(Timetable).filter(Timetable.id == 1).first()
        if current:
            current.data = timetable
        else:
            # Force insert with id=1
            from sqlalchemy import text
            db.execute(text("INSERT OR REPLACE INTO timetable (id, data) VALUES (1, :data)"),
                       {"data": __import__('json').dumps(timetable)})
        db.commit()
    except Exception as e:
        print(f"Save error: {e}")
        # Fallback: just try a simple save
        try:
            db.rollback()
            current = db.query(Timetable).filter(Timetable.id == 1).first()
            if current:
                current.data = timetable
                db.commit()
            else:
                new_tt = Timetable(data=timetable)
                db.add(new_tt)
                db.commit()
        except Exception as e2:
            print(f"Fallback save error: {e2}")

    return {"ok": True, "timetable": timetable}


@app.get("/api/timetable/history")
def get_timetable_history(db: Session = Depends(get_db)):
    all_tt = db.query(Timetable).filter(Timetable.id != 1).order_by(Timetable.id.desc()).all()
    return [{"id": t.id, "saved_at": t.data.get("saved_at", "Unknown") if isinstance(t.data, dict) and "saved_at" in t.data else "Unknown"} for t in all_tt]


@app.delete("/api/timetable/history")
def clear_timetable_history(db: Session = Depends(get_db)):
    deleted = db.query(Timetable).filter(Timetable.id != 1).delete()
    db.commit()
    return {"ok": True, "deleted": deleted}


@app.get("/api/timetable/history/{history_id}")
def get_timetable_version(history_id: int, db: Session = Depends(get_db)):
    tt = db.query(Timetable).filter(Timetable.id == history_id).first()
    if not tt:
        raise HTTPException(404, "Not found")
    data = tt.data
    if isinstance(data, dict) and "timetable" in data:
        return data["timetable"]
    return data


@app.post("/api/timetable/restore/{history_id}")
def restore_timetable(history_id: int, db: Session = Depends(get_db)):
    tt = db.query(Timetable).filter(Timetable.id == history_id).first()
    if not tt:
        raise HTTPException(404, "Not found")
    data = tt.data
    timetable_data = data.get("timetable", data) if isinstance(data, dict) else data
    current = db.query(Timetable).filter(Timetable.id == 1).first()
    if current:
        current.data = timetable_data
    else:
        db.add(Timetable(data=timetable_data))
    db.commit()
    return {"ok": True}


@app.post("/api/timetable/publish")
def publish_timetable(request: Request, db: Session = Depends(get_db)):
    """Publish current timetable - only admin can do this"""
    role = get_current_role(request)
    if role != "admin":
        raise HTTPException(403, "Only admin can publish timetables")

    # Get current timetable
    tt = db.query(Timetable).filter(Timetable.id == 1).first()
    if not tt:
        # Fallback to latest
        tt = db.query(Timetable).order_by(Timetable.id.desc()).first()
    if not tt:
        raise HTTPException(404, "No timetable to publish")

    data = tt.data
    if isinstance(data, dict) and "timetable" in data:
        data = data["timetable"]

    # Save as published (id=2 is reserved for published version)
    from datetime import datetime
    published = db.query(Timetable).filter(Timetable.id == 2).first()
    if published:
        published.data = {"timetable": data, "published_at": datetime.utcnow().isoformat()}
    else:
        from sqlalchemy import text
        db.execute(text("INSERT INTO timetable (id, data) VALUES (2, :data)"),
                   {"data": __import__('json').dumps({"timetable": data, "published_at": datetime.utcnow().isoformat()})})
    db.commit()
    return {"ok": True, "message": "Timetable published successfully"}


@app.get("/api/timetable/published")
def get_published_timetable(db: Session = Depends(get_db)):
    """Get the published timetable - accessible by staff"""
    tt = db.query(Timetable).filter(Timetable.id == 2).first()
    if not tt:
        return {}
    data = tt.data
    if isinstance(data, dict) and "timetable" in data:
        return data["timetable"]
    return data if data else {}


@app.get("/api/classes/export-excel")
def export_classes_excel(db: Session = Depends(get_db)):
    """Export all class data as Excel file"""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    from fastapi.responses import StreamingResponse

    classes = db.query(SchoolClass).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classes"

    # Headers
    headers = ['Class', 'Division', 'Block', 'Type', 'Class Teacher', 'Subject', 'Teacher', 'Periods/Week']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    row = 2
    for cls in classes:
        first_sub = True
        for sub in cls.subjects or []:
            if sub.get('periodsPerWeek', 0) <= 0:
                continue
            if first_sub:
                ws.cell(row=row, column=1, value=cls.name)
                ws.cell(row=row, column=2, value=cls.divisions[0] if cls.divisions else '')
                ws.cell(row=row, column=3, value=cls.block or '')
                ws.cell(row=row, column=4, value=cls.classType or '')
                ws.cell(row=row, column=5, value=cls.classTeacher or '')
                first_sub = False
            ws.cell(row=row, column=6, value=sub.get('name', ''))
            ws.cell(row=row, column=7, value=sub.get('teacher', ''))
            ws.cell(row=row, column=8, value=sub.get('periodsPerWeek', 0))
            row += 1
        row += 1  # blank row between classes

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=classes_export.xlsx"}
    )


# ===== Serve Frontend =====
frontend_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")


@app.get("/")
def serve_login():
    return FileResponse(os.path.join(frontend_path, "login.html"))


@app.get("/app")
def serve_app():
    return FileResponse(os.path.join(frontend_path, "index.html"))


@app.get("/styles.css")
def serve_css():
    return FileResponse(os.path.join(frontend_path, "styles.css"))


@app.get("/app.js")
def serve_js():
    return FileResponse(os.path.join(frontend_path, "app.js"))
