"""Creates an Excel template for bulk class data entry - Division-wise teacher assignment."""
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
class_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

subjects_8_9 = [
    'First Language', 'Malayalam II', 'English', 'Hindi', 'Maths',
    'Social Science', 'Chemistry', 'Physics', 'Biology', 'IT',
    'PET', 'Music', 'Work Education', 'Art'
]

subjects_10 = [
    'First Language', 'Malayalam II', 'English', 'Hindi', 'Maths',
    'Social Science', 'Chemistry', 'Physics', 'Biology', 'IT'
]

# ===== CLASSES SHEET =====
ws = wb.active
ws.title = "Classes"

# Title
ws.merge_cells('A1:G1')
ws['A1'] = "KKHMS Alathiyur - Class Data Entry (Division-wise Teacher Assignment)"
ws['A1'].font = Font(bold=True, size=13, color="4F46E5")

ws.merge_cells('A2:G2')
ws['A2'] = "Each row = one DIVISION. Fill Class, Division, Block, Type, Class Teacher, then subjects with teacher & periods."
ws['A2'].font = Font(italic=True, size=10, color="666666")

# Headers (Row 4)
headers = ['Class', 'Division', 'Block', 'Type', 'Class Teacher', 'Subject', 'Teacher', 'Periods/Week']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

widths = [8, 10, 15, 25, 20, 20, 20, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Example: Class 8 with divisions A, B, C
row = 5
divisions_8 = ['A', 'B', 'C', 'D', 'E', 'F']  # Example 6 divisions

for div_idx, div in enumerate(divisions_8):
    for sub_idx, sub in enumerate(subjects_8_9):
        r = row
        cell_class = ws.cell(row=r, column=1)
        cell_div = ws.cell(row=r, column=2)
        cell_block = ws.cell(row=r, column=3)
        cell_type = ws.cell(row=r, column=4)
        cell_ct = ws.cell(row=r, column=5)

        if sub_idx == 0:
            cell_class.value = "8"
            cell_div.value = div
            cell_block.value = "Block A"
            cell_type.value = "Arabic"
            cell_ct.value = ""
            # Highlight first row of each division
            for c in range(1, 9):
                ws.cell(row=r, column=c).fill = class_fill

        ws.cell(row=r, column=6, value=sub)
        ws.cell(row=r, column=7, value="")  # Teacher to fill
        ws.cell(row=r, column=8, value=1 if sub in ['PET', 'Music', 'Work Education', 'Art'] else "")

        for c in range(1, 9):
            ws.cell(row=r, column=c).border = thin_border

        row += 1

    # Add empty row between divisions for readability
    row += 1

# Add Class 9 example with 3 divisions
ws.cell(row=row, column=1, value="--- Class 9 ---").font = Font(bold=True, color="4F46E5")
row += 1

divisions_9 = ['A', 'B', 'C']
for div in divisions_9:
    for sub_idx, sub in enumerate(subjects_8_9):
        r = row
        if sub_idx == 0:
            ws.cell(row=r, column=1, value="9")
            ws.cell(row=r, column=2, value=div)
            ws.cell(row=r, column=3, value="Block A")
            ws.cell(row=r, column=4, value="")
            ws.cell(row=r, column=5, value="")
            for c in range(1, 9):
                ws.cell(row=r, column=c).fill = class_fill

        ws.cell(row=r, column=6, value=sub)
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=8, value=1 if sub in ['PET', 'Music', 'Work Education', 'Art'] else "")

        for c in range(1, 9):
            ws.cell(row=r, column=c).border = thin_border
        row += 1
    row += 1

# Add Class 10 example with 2 divisions
ws.cell(row=row, column=1, value="--- Class 10 ---").font = Font(bold=True, color="4F46E5")
row += 1

divisions_10 = ['A', 'B']
for div in divisions_10:
    for sub_idx, sub in enumerate(subjects_10):
        r = row
        if sub_idx == 0:
            ws.cell(row=r, column=1, value="10")
            ws.cell(row=r, column=2, value=div)
            ws.cell(row=r, column=3, value="Block B")
            ws.cell(row=r, column=4, value="")
            ws.cell(row=r, column=5, value="")
            for c in range(1, 9):
                ws.cell(row=r, column=c).fill = class_fill

        ws.cell(row=r, column=6, value=sub)
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=8, value="")

        for c in range(1, 9):
            ws.cell(row=r, column=c).border = thin_border
        row += 1
    row += 1


# ===== INSTRUCTIONS SHEET =====
ws2 = wb.create_sheet("Instructions")
instructions = [
    "HOW TO FILL THIS TEMPLATE",
    "",
    "STRUCTURE: Each ROW is one subject for one division.",
    "A new division starts whenever Column A (Class) AND Column B (Division) have values.",
    "",
    "COLUMNS:",
    "  A - Class: 8, 9, or 10 (fill only on first subject row of each division)",
    "  B - Division: A, B, C, D... (fill only on first subject row of each division)",
    "  C - Block: Block name as in app (fill only on first row of each division)",
    "  D - Type: Arabic/Malayalam/etc. (fill only on first row of each division)",
    "  E - Class Teacher: Teacher name (fill only on first row of each division)",
    "  F - Subject: Pre-filled subject name",
    "  G - Teacher: The teacher who teaches THIS subject to THIS division",
    "  H - Periods/Week: Number of periods per week",
    "",
    "EXAMPLE:",
    "  Class | Div | Block   | Type   | CT     | Subject    | Teacher      | Periods",
    "  8     | A   | Block A | Arabic | Salma  | English    | Suresh Kumar | 5",
    "        |     |         |        |        | Maths      | Priya R      | 5",
    "        |     |         |        |        | Hindi      | Fathima N    | 4",
    "  8     | B   | Block A | Arabic | Noor   | English    | Divya M      | 5",
    "        |     |         |        |        | Maths      | Rajesh M     | 5",
    "        |     |         |        |        | Hindi      | Fathima N    | 4",
    "",
    "RULES:",
    "- Total periods per division MUST equal 35 (5 days × 7 periods)",
    "- PET, Music, Work Education, Art: fixed at 1 period/week (Class 8 & 9 only)",
    "- Class 10 does NOT have PET, Music, WE, Art",
    "- Teacher names must EXACTLY match names in the Teachers list",
    "- You can have different teachers for same subject in different divisions",
    "- Add/remove division rows as needed (copy a set of subject rows)",
    "",
    "TYPES AVAILABLE:",
    "  Arabic, Malayalam, Urdu/Sanskrit, Sanskrit/Urdu/Arabic,",
    "  Malayalam/Arabic, Sanskrit/Arabic/Urdu/Malayalam, Sanskrit/Arabic,",
    "  Urdu/Arabic, Sanskrit, Urdu",
    "",
    "After filling, upload this file in the app: Classes page → 'Upload Excel' button."
]

for i, line in enumerate(instructions, 1):
    ws2.cell(row=i, column=1, value=line)
    if i == 1:
        ws2.cell(row=i, column=1).font = Font(bold=True, size=14)
    elif line.startswith("COLUMNS:") or line.startswith("RULES:") or line.startswith("TYPES") or line.startswith("EXAMPLE:"):
        ws2.cell(row=i, column=1).font = Font(bold=True)

ws2.column_dimensions['A'].width = 90

# Save
output_path = r"c:\Users\mushkott\Desktop\Python Projects\school-timetable\class_template.xlsx"
wb.save(output_path)
print(f"✅ Template created: {output_path}")
print(f"   - Class 8: {len(divisions_8)} divisions × {len(subjects_8_9)} subjects")
print(f"   - Class 9: {len(divisions_9)} divisions × {len(subjects_8_9)} subjects")
print(f"   - Class 10: {len(divisions_10)} divisions × {len(subjects_10)} subjects")
